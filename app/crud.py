from sqlalchemy import select, MetaData, Table, Column, Integer, String, Float, inspect, text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from fastapi.responses import StreamingResponse
from . import models, schemas
from .database import engine
import logging

logger = logging.getLogger(__name__)

#commit
async def get_test(db: AsyncSession, test_id: str):
    result = await db.execute(select(models.Test).where(models.Test.test_id == test_id))
    return result.scalars().first()

async def telegram_id_exists(db: AsyncSession, table_name: str, telegram_id: str) -> bool:
    """Return False if telegram_id exists in the given table, else True."""
    try:
        query = text(f"""
            SELECT EXISTS (
                SELECT 1 FROM {table_name} WHERE telegram_id = :telegram_id
            );
        """)
        result = await db.execute(query, {"telegram_id": telegram_id})
        exists = result.scalar()
        return not exists  # Invert: if exists -> False, if not -> True
    except SQLAlchemyError as e:
        logger.error(f"Error checking telegram_id in table '{table_name}': {str(e)}")
        return False


async def create_test(db: AsyncSession, test: schemas.TestCreate):
    db_test = models.Test(**test.dict())
    db.add(db_test)
    await db.commit()
    await db.refresh(db_test)
    return db_test


async def table_exists(sync_conn, table_name: str) -> bool:
    """Check if table exists using synchronous connection"""
    inspector = inspect(sync_conn)
    return table_name in inspector.get_table_names()


async def table_exists(db: AsyncSession, table_name: str) -> bool:
    """Check if table exists using async session"""
    try:
        # Use raw SQL with text() to check table existence
        result = await db.execute(
            text(
                """
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_name = :table_name
                );
                """
            ).bindparams(table_name=table_name)
        )
        return result.scalar()
    except SQLAlchemyError as e:
        logger.error(f"Error checking table existence: {str(e)}")
        return False


async def create_user_response(db: AsyncSession, response: schemas.UserResponseCreate):
    try:
        # Get the test to validate answers
        test = await get_test(db, response.test_id)
        if not test:
            logger.error(f"Test not found: {response.test_id}")
            return None

        # Validate answer length
        if len(response.answers) != len(test.answers):
            logger.error(f"Answer length mismatch: {len(response.answers)} vs {len(test.answers)}")
            return None

        # Calculate score
        score = 0.0
        for i, (user_answer, correct_answer) in enumerate(zip(response.answers, test.answers)):
            if i < 45:
                points = 1.1
            elif i < 75:  # 45 + 30
                points = 3.1
            else:  # remaining 30
                points = 2.1

            if user_answer == correct_answer:
                score += points

        # Table name (sanitized)
        table_name = f"{response.test_id}_answers"

        # Check if table exists
        exists = await table_exists(db, table_name)

        if not exists:
            # Create new table using text() for raw SQL
            await db.execute(
                text(
                    f"""
                    CREATE TABLE IF NOT EXISTS {table_name} (
                        id SERIAL PRIMARY KEY,
                        telegram_id VARCHAR(20) NOT NULL,
                        first_name VARCHAR NOT NULL,
                        last_name VARCHAR NOT NULL,
                        middle_name VARCHAR,
                        region VARCHAR NOT NULL,
                        answers VARCHAR NOT NULL,
                        score FLOAT NOT NULL
                    );
                    """
                )
            )
            await db.commit()

        # Insert response data using parameterized query
        result = await db.execute(
            text(
                f"""
                INSERT INTO {table_name} 
                (telegram_id ,first_name, last_name, middle_name, region, answers, score)
                VALUES (:telegram_id, :first_name, :last_name, :middle_name, :region, :answers, :score)
                RETURNING id;
                """
            ).bindparams(
                telegram_id=response.telegram_id,
                first_name=response.first_name,
                last_name=response.last_name,
                middle_name=response.middle_name,
                region=response.region,
                answers=response.answers,
                score=round(score, 2)
            )
        )
        await db.commit()

        inserted_id = result.scalar()

        return {
            "id": inserted_id,
            **response.dict(),
            "score": round(score, 2)
        }

    except Exception as e:
        logger.error(f"Error creating user response: {str(e)}")
        await db.rollback()
        raise

async def get_all_tests(db: AsyncSession, skip: int = 0, limit: int = 100):
    """Barcha testlarni paginatsiya bilan qaytaradi"""
    try:
        result = await db.execute(
            select(models.Test)
            .offset(skip)
            .limit(limit)
            .order_by(models.Test.id)
        )
        return result.scalars().all()
    except Exception as e:
        logger.error(f"Error getting all tests: {str(e)}")
        raise

async def delete_test(db: AsyncSession, test_id: str):
    """Testni va unga tegishli javoblar jadvalini o'chiradi"""
    try:
        # 1. Testni topamiz
        test = await get_test(db, test_id)
        if not test:
            return False

        # 2. Testni o'chiramiz
        await db.delete(test)

        # 3. Test javoblari jadvalini o'chiramiz
        table_name = f"{test_id}_responses"
        if await table_exists(db, table_name):
            await db.execute(
                text(f"DROP TABLE {table_name};")
            )

        await db.commit()
        return True

    except Exception as e:
        logger.error(f"Error deleting test {test_id}: {str(e)}")
        await db.rollback()
        return False

async def get_user_responses(db: AsyncSession, test_id: str, skip: int = 0, limit: int = 100):
    try:
        table_name = f"{test_id}_answers"

        # Check if table exists
        exists = await table_exists(db, table_name)

        if not exists:
            return []

        # Get responses using parameterized query
        result = await db.execute(
            text(
                f"""
                SELECT * FROM {table_name}
                ORDER BY id
                OFFSET :skip LIMIT :limit;
                """
            ).bindparams(skip=skip, limit=limit)
        )

        return [dict(row) for row in result.mappings()]

    except Exception as e:
        logger.error(f"Error getting user responses: {str(e)}")
        raise

async def export_test_results_to_excel(db: AsyncSession, test_id: str):
    """Export test results to Excel file with specified format"""
    try:
        # Check if test exists
        test = await get_test(db, test_id)
        if not test:
            return None

        # Get responses from dynamic table
        table_name = f"{test_id}_answers"
        if not await table_exists(db, table_name):
            return None

        # Get data ordered by score descending
        result = await db.execute(
            text(
                f"""
                SELECT first_name, last_name, region, score 
                FROM {table_name}
                ORDER BY score DESC;
                """
            )
        )
        responses = result.mappings().all()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Add headers
        headers = ["No", "F.I.O (Region)", "Ball"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Add data rows
        for idx, row in enumerate(responses, start=1):
            fio = f"{row['first_name']} {row['last_name']} ({row['region']})"
            ws.append([idx, fio, row['score']])

        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10

        # Center the No and Ball columns
        for row in ws.iter_rows(min_row=2):
            row[0].alignment = Alignment(horizontal='center')  # No
            row[2].alignment = Alignment(horizontal='center')  # Ball

        # Save to bytes buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    except Exception as e:
        logger.error(f"Error exporting test results: {str(e)}")
        return None